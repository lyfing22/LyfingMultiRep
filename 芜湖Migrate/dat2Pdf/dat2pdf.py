"""
dat2pdf — 批量 dat → jpg → pdf 转换器
- 遍历根目录下形如 20230522 的日期文件夹
- 递归处理 .dat 文件
- 步骤1: 调用 ConvertReport2Image.exe 转 jpg
- 步骤2: img2pdf 转 pdf（保留原像素尺寸、无损）
- 失败记录到 conversion_failures.xlsx（1万行自动归档）
- ROOT / EXE 路径可配置（CLI > 环境变量 > config.json > 默认）
"""
import os
import sys
import json
import subprocess
from datetime import datetime
from pathlib import Path
import img2pdf
from openpyxl import load_workbook, Workbook

def _script_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent.resolve()
    return Path(__file__).parent.resolve()

SCRIPT_DIR     = _script_dir()
CONFIG_PATH    = SCRIPT_DIR / "config.json"
EXCEL          = SCRIPT_DIR / "conversion_failures.xlsx"
EXCEL_PATTERN  = "conversion_failures_*.xlsx"

DEFAULT_CONFIG = {
    "root":        r"D:\Migrate\Report_Data",
    "exe_dir":     "dat2Img",
    "exe_name":    "ConvertReport2Image.exe",
    "max_rows":    10_000,
    "timeout_sec": 120,
}
DAT_EXTS  = {".dat"}
DATE_LEN  = 8
EXE_DIR_KEYWORD = "dat2Img"
HEADER = ["dat_path","date_folder","stage","error_msg",
          "attempt_count","first_seen","last_seen"]


def load_config() -> dict:
    cfg = DEFAULT_CONFIG.copy()

    if CONFIG_PATH.exists():
        try:
            cfg.update(json.loads(CONFIG_PATH.read_text(encoding="utf-8")))
        except Exception as e:
            print(f"[警告] config.json 解析失败: {e}  使用默认配置")

    env_root = os.environ.get("DAT2PDF_ROOT")
    if env_root:
        cfg["root"] = env_root

    if len(sys.argv) > 1:
        cfg["root"] = sys.argv[1]

    return cfg


def resolve_paths(cfg: dict) -> dict:
    root = Path(cfg["root"]).resolve()
    exe_dir = Path(cfg["exe_dir"])
    if not exe_dir.is_absolute():
        exe_dir = SCRIPT_DIR / exe_dir
    exe = exe_dir / cfg["exe_name"]
    return {
        "root":        root,
        "exe":         exe,
        "max_rows":    int(cfg["max_rows"]),
        "timeout_sec": int(cfg["timeout_sec"]),
    }


def ensure_config():
    if not CONFIG_PATH.exists():
        CONFIG_PATH.write_text(
            json.dumps(DEFAULT_CONFIG, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        print(f"[信息] 已生成配置文件: {CONFIG_PATH}")


def _archive_current():
    if not EXCEL.exists():
        return
    used = set()
    for p in EXCEL.parent.glob(EXCEL_PATTERN):
        if p == EXCEL:
            continue
        suffix = p.stem.split("_")[-1]
        if suffix.isdigit():
            used.add(int(suffix))
    nxt = (max(used) + 1) if used else 1
    EXCEL.rename(EXCEL.with_name(f"conversion_failures_{nxt:03d}.xlsx"))
    print(f"[归档] {EXCEL.name} -> conversion_failures_{nxt:03d}.xlsx")


def _ensure_file(max_rows: int):
    if EXCEL.exists():
        wb = load_workbook(EXCEL)
        ws = wb["failures"]
        if ws.max_row - 1 >= max_rows:
            wb.close()
            _archive_current()
            wb = Workbook()
            ws = wb.active
            ws.title = "failures"
            ws.append(HEADER)
            wb.save(EXCEL)
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "failures"
    ws.append(HEADER)
    wb.save(EXCEL)
    return wb, ws


def log_failure(dat: Path, date_folder: str, stage: str, err: str, max_rows: int):
    wb, ws = _ensure_file(max_rows)
    key = str(dat)
    now = datetime.now().isoformat(timespec="seconds")
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == key:
            row[2].value, row[3].value = stage, str(err)
            row[4].value = (row[4].value or 0) + 1
            row[6].value = now
            wb.save(EXCEL)
            return
    ws.append([key, date_folder, stage, str(err), 1, now, now])
    wb.save(EXCEL)


def _sorted_pages(dir: Path, stem: str) -> list[Path]:
    return sorted(
        dir.glob(f"{stem}_*.jpg"),
        key=lambda p: int(p.stem.rsplit("_", 1)[-1]),
    )


def dat_to_jpg(dat: Path, exe: Path, exe_dir: Path, timeout: int) -> list[Path]:
    stem = dat.stem
    first = dat.parent / f"{stem}_0.jpg"

    if first.exists():
        return _sorted_pages(dat.parent, stem)

    subprocess.run(
        [str(exe), str(dat), str(dat.parent)],
        cwd=str(exe_dir), check=True, timeout=timeout,
    )

    if not first.exists():
        raise FileNotFoundError(f"EXE 未生成 JPG: 期望 {first.name}")

    return _sorted_pages(dat.parent, stem)


def jpg_to_pdf(jpg_list: list[Path]) -> Path:
    pdf = jpg_list[0].with_suffix(".pdf")
    pdf.write_bytes(img2pdf.convert([str(p) for p in jpg_list]))
    return pdf


def is_date_folder(name: str) -> bool:
    return len(name) == DATE_LEN and name.isdigit()


def process_date_dir(date_dir: Path, paths: dict):
    for dat in date_dir.rglob("*"):
        if not dat.is_file() or dat.suffix.lower() not in DAT_EXTS:
            continue
        if EXE_DIR_KEYWORD in dat.parts:
            continue

        pdf = dat.with_suffix(".pdf")
        first_jpg = dat.parent / f"{dat.stem}_0.jpg"

        if pdf.exists():
            print(f"  [跳过] {dat.relative_to(date_dir)}  (PDF 已存在)")
            continue
        if first_jpg.exists():
            print(f"  [跳过] {dat.relative_to(date_dir)}  (JPG 已存在)")

        try:
            jpg_list = dat_to_jpg(dat, paths["exe"], paths["exe"].parent,
                                  paths["timeout_sec"])
            if not jpg_list:
                raise FileNotFoundError("EXE 未生成 JPG")
            pdf = jpg_to_pdf(jpg_list)
            print(f"  [完成] {dat.name} -> {pdf.name}")
        except FileNotFoundError as e:
            log_failure(dat, date_dir.name, "jpg_missing", str(e),
                        paths["max_rows"])
            print(f"  [失败] {dat.name}  {e}")
        except subprocess.CalledProcessError as e:
            log_failure(dat, date_dir.name, "exe_call",
                        f"EXE exit {e.returncode}", paths["max_rows"])
            print(f"  [失败] {dat.name}  EXE exit {e.returncode}")
        except OSError as e:
            log_failure(dat, date_dir.name, "exe_launch", str(e),
                        paths["max_rows"])
            print(f"  [失败] {dat.name}  EXE 启动失败: {e}")
        except Exception as e:
            log_failure(dat, date_dir.name, "jpg_to_pdf", str(e),
                        paths["max_rows"])
            print(f"  [失败] {dat.name}  {e}")


def main():
    ensure_config()
    cfg = load_config()
    paths = resolve_paths(cfg)

    if not paths["root"].exists():
        print(f"[错误] 根目录不存在: {paths['root']}"); sys.exit(1)
    if not paths["exe"].exists():
        print(f"[错误] EXE 不存在: {paths['exe']}"); sys.exit(1)

    print(f"[开始] 根目录: {paths['root']}")
    print(f"       EXE:    {paths['exe']}")
    print(f"       日志:   {EXCEL}\n")

    for d in sorted(paths["root"].iterdir()):
        if d.is_dir() and is_date_folder(d.name):
            print(f"[目录] {d.name}/")
            process_date_dir(d, paths)
    print(f"\n[完成] 失败记录: {EXCEL}")


if __name__ == "__main__":
    main()
